# ============================================================
# INVOICE OCR SYSTEM
# EasyOCR + LayoutLMv3 + 2D OCR
#
# FEATURES
# ------------------------------------------------------------
# 1. Single invoice image
# 2. Multiple invoice images
# 3. Folder upload
# 4. Batch processing
# 5. Retry OCR with enhanced image
# 6. Continues when one invoice fails
# 7. Excel: New / Continue Existing
# 8. Keeps all processed invoices
# 9. Does not store 120 full images in session state
# 10. Progress + success/failure report
# ============================================================


# ============================================================
# SSL CERTIFICATE FIX
# ============================================================

import os
import ssl
import certifi

os.environ["SSL_CERT_FILE"] = certifi.where()
os.environ["REQUESTS_CA_BUNDLE"] = certifi.where()

ssl._create_default_https_context = (
    lambda: ssl.create_default_context(
        cafile=certifi.where()
    )
)


# ============================================================
# IMPORTS
# ============================================================

import re
import io
import gc
from datetime import datetime

import streamlit as st
import easyocr
import numpy as np
import torch
import pandas as pd

from PIL import Image, ImageEnhance

from transformers import (
    AutoTokenizer,
    LayoutLMv3ForTokenClassification
)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# PAGE CONFIG
# ============================================================

st.set_page_config(
    page_title="Invoice OCR System",
    page_icon="📄",
    layout="wide"
)


# ============================================================
# MODEL PATH
# ============================================================

MODEL_PATH = (
    r"C:\Users\tatak\haroon\Invoice OCR System"
    r"\saved_model\final_model"
)


# ============================================================
# DEVICE
# ============================================================

DEVICE = torch.device(
    "cuda"
    if torch.cuda.is_available()
    else "cpu"
)


# ============================================================
# BATCH SIZE
#
# IMPORTANT:
# This does NOT mean only this many invoices are processed.
# It controls how many uploaded files are handled before
# memory cleanup.
# ============================================================

BATCH_SIZE = 5


# ============================================================
# SESSION STATE
# ============================================================

if "excel_rows" not in st.session_state:
    st.session_state.excel_rows = []

if "loaded_excel_name" not in st.session_state:
    st.session_state.loaded_excel_name = None

if "processing_results" not in st.session_state:
    st.session_state.processing_results = []

if "failed_files" not in st.session_state:
    st.session_state.failed_files = []

if "processed_file_names" not in st.session_state:
    st.session_state.processed_file_names = set()

if "processing_complete" not in st.session_state:
    st.session_state.processing_complete = False


# ============================================================
# LOAD LAYOUTLMV3
# ============================================================

@st.cache_resource
def load_layoutlmv3():

    tokenizer = AutoTokenizer.from_pretrained(
        MODEL_PATH,
        local_files_only=True
    )

    model = LayoutLMv3ForTokenClassification.from_pretrained(
        MODEL_PATH,
        local_files_only=True
    )

    model.to(DEVICE)

    model.eval()

    return tokenizer, model


# ============================================================
# LOAD EASYOCR
# ============================================================

@st.cache_resource
def load_ocr():

    reader = easyocr.Reader(
        ["en"],
        gpu=torch.cuda.is_available(),
        verbose=False
    )

    return reader


# ============================================================
# LOAD MODELS
# ============================================================

try:

    with st.spinner("Loading OCR models..."):

        reader = load_ocr()

        tokenizer, model = load_layoutlmv3()

except Exception as e:

    st.error(
        "Could not load the OCR models."
    )

    st.exception(e)

    st.stop()


# ============================================================
# CLEAN TEXT
# ============================================================

def clean_text(value):

    if value is None:
        return ""

    value = str(value)

    value = value.replace(
        "\n",
        " "
    )

    value = re.sub(
        r"\s+",
        " ",
        value
    )

    return value.strip()


# ============================================================
# CLEAN MONEY
# ============================================================

def clean_money(value):

    if not value:
        return ""

    value = str(value)

    value = value.replace(
        "S",
        "$"
    )

    value = value.replace(
        "s",
        "$"
    )

    value = value.replace(
        "O",
        "0"
    )

    value = value.replace(
        "o",
        "0"
    )

    value = re.sub(
        r"[^0-9$.,\-]",
        "",
        value
    )

    return value.strip()


# ============================================================
# OCR FUNCTION
# ============================================================

def run_easyocr(
    image,
    reader,
    scale=3
):

    # --------------------------------------------------------
    # RESIZE
    # --------------------------------------------------------

    width = int(
        image.width * scale
    )

    height = int(
        image.height * scale
    )

    ocr_image = image.resize(
        (
            width,
            height
        ),
        Image.Resampling.LANCZOS
    )

    # --------------------------------------------------------
    # CONTRAST
    # --------------------------------------------------------

    ocr_image = ImageEnhance.Contrast(
        ocr_image
    ).enhance(1.2)

    image_array = np.array(
        ocr_image
    )

    # --------------------------------------------------------
    # OCR
    # --------------------------------------------------------

    results = reader.readtext(
        image_array,
        detail=1
    )

    words = []
    boxes = []

    for result in results:

        if len(result) < 2:
            continue

        polygon = result[0]

        text = clean_text(
            result[1]
        )

        if not text:
            continue

        xs = [
            point[0]
            for point in polygon
        ]

        ys = [
            point[1]
            for point in polygon
        ]

        x1 = min(xs)
        y1 = min(ys)

        x2 = max(xs)
        y2 = max(ys)

        words.append(
            text
        )

        boxes.append(
            [
                float(x1),
                float(y1),
                float(x2),
                float(y2)
            ]
        )

    return words, boxes


# ============================================================
# OCR WITH RETRY
#
# First attempt = normal
# Second attempt = stronger preprocessing
# ============================================================

def extract_ocr(
    image,
    reader
):

    # --------------------------------------------------------
    # ATTEMPT 1
    # --------------------------------------------------------

    words, boxes = run_easyocr(
        image,
        reader,
        scale=3
    )

    if words:

        return words, boxes

    # --------------------------------------------------------
    # ATTEMPT 2
    # --------------------------------------------------------

    enhanced = ImageEnhance.Contrast(
        image
    ).enhance(1.6)

    enhanced = ImageEnhance.Sharpness(
        enhanced
    ).enhance(1.5)

    words, boxes = run_easyocr(
        enhanced,
        reader,
        scale=4
    )

    del enhanced

    gc.collect()

    return words, boxes


# ============================================================
# CREATE VISUAL LINES
# ============================================================

def create_visual_lines(
    words,
    boxes
):

    if len(words) != len(boxes):

        raise ValueError(
            f"Words: {len(words)}, "
            f"Boxes: {len(boxes)}"
        )

    items = []

    for word, box in zip(
        words,
        boxes
    ):

        word = clean_text(
            word
        )

        if not word:
            continue

        x1, y1, x2, y2 = box

        items.append({

            "word": word,

            "x1": x1,
            "y1": y1,

            "x2": x2,
            "y2": y2,

            "cx":
                (x1 + x2) / 2,

            "cy":
                (y1 + y2) / 2,

            "height":
                max(
                    y2 - y1,
                    1
                )

        })

    if not items:

        return []

    items.sort(
        key=lambda x: (
            x["cy"],
            x["x1"]
        )
    )

    lines = []

    for item in items:

        placed = False

        for line in lines:

            line_top = min(
                x["y1"]
                for x in line
            )

            line_bottom = max(
                x["y2"]
                for x in line
            )

            line_height = max(
                line_bottom - line_top,
                1
            )

            overlap = (

                min(
                    item["y2"],
                    line_bottom
                )

                -

                max(
                    item["y1"],
                    line_top
                )

            )

            minimum_height = min(
                item["height"],
                line_height
            )

            if overlap >= (
                0.35 *
                minimum_height
            ):

                line.append(
                    item
                )

                placed = True

                break

        if not placed:

            lines.append(
                [item]
            )

    for line in lines:

        line.sort(
            key=lambda x:
            x["x1"]
        )

    lines.sort(

        key=lambda line:
        min(
            x["y1"]
            for x in line
        )

    )

    return lines


# ============================================================
# CREATE 2D ALIGNED OCR
# ============================================================

def create_aligned_ocr(
    lines
):

    if not lines:
        return ""

    items = []

    for line in lines:

        items.extend(
            line
        )

    if not items:
        return ""

    ocr_width = max(
        item["x2"]
        for item in items
    )

    if ocr_width <= 0:
        return ""

    CONSOLE_WIDTH = 100

    scale_x = (
        CONSOLE_WIDTH /
        ocr_width
    )

    aligned_lines = []

    for line in lines:

        output = ""

        current_column = 0

        for item in line:

            target_column = int(
                item["x1"] *
                scale_x
            )

            spaces = (
                target_column -
                current_column
            )

            if spaces < 1:

                spaces = 1

            output += (
                " " * spaces
            )

            output += (
                item["word"]
            )

            current_column = (
                target_column
                +
                len(item["word"])
            )

        if output.strip():

            aligned_lines.append(
                output.rstrip()
            )

    return "\n".join(
        aligned_lines
    )


# ============================================================
# FIND LABEL VALUE
# ============================================================

def find_label_value(
    text,
    labels
):

    lines = text.splitlines()

    for line in lines:

        line = clean_text(
            line
        )

        for label in labels:

            pattern = (

                r"(?i)"
                +
                re.escape(label)
                +
                r"\s*[:#]?\s*(.+)"

            )

            match = re.search(
                pattern,
                line
            )

            if match:

                value = clean_text(
                    match.group(1)
                )

                if value:

                    return value

    return ""


# ============================================================
# FIND MONEY
# ============================================================

def find_money_after_label(
    text,
    labels
):

    lines = text.splitlines()

    for line in lines:

        line = clean_text(
            line
        )

        for label in labels:

            pattern = (

                r"(?i)"
                +
                re.escape(label)
                +
                r"\s*[:#]?\s*"
                r"([$S]?\s*"
                r"[\d,]+"
                r"(?:\.\d{1,2})?)"

            )

            match = re.search(
                pattern,
                line
            )

            if match:

                return clean_money(
                    match.group(1)
                )

    return ""


# ============================================================
# INVOICE NUMBER
# ============================================================

def extract_invoice_number(
    text
):

    patterns = [

        r"(?i)invoice\s*(?:number|no\.?|#)?\s*[:#]?\s*([A-Z0-9\-_\/]+)",

        r"(?i)PO\s*[-:#]?\s*([A-Z0-9\-_\/]+)",

        r"(?i)INV\s*[-:#]?\s*([A-Z0-9\-_\/]+)"

    ]

    for pattern in patterns:

        match = re.search(
            pattern,
            text
        )

        if match:

            return clean_text(
                match.group(1)
            )

    return ""


# ============================================================
# COMPANY
# ============================================================

def extract_company(
    text
):

    lines = [

        clean_text(line)

        for line
        in text.splitlines()

        if clean_text(line)

    ]

    ignore = [

        "invoice",
        "bill to",
        "description",
        "amount",
        "total",
        "subtotal",
        "tax",
        "paid",
        "balance"

    ]

    for line in lines[:10]:

        lower = line.lower()

        if any(
            word in lower
            for word in ignore
        ):

            continue

        if len(line) >= 3:

            return line

    return ""


# ============================================================
# CUSTOMER
# ============================================================

def extract_customer(
    text
):

    lines = [

        clean_text(line)

        for line
        in text.splitlines()

        if clean_text(line)

    ]

    for index, line in enumerate(
        lines
    ):

        if "bill to" in line.lower():

            value = re.sub(

                r"(?i).*bill\s*to\s*[:#]?",

                "",

                line

            )

            value = clean_text(
                value
            )

            value = re.split(

                r"(?i)invoice\s*(number|no\.?|#)?",

                value

            )[0]

            customer_name = clean_text(
                value
            )

            address_lines = []

            for next_line in lines[
                index + 1:
                index + 5
            ]:

                lower = next_line.lower()

                if any(
                    word in lower
                    for word in [
                        "description",
                        "invoice number",
                        "date",
                        "term",
                        "due"
                    ]
                ):

                    break

                address_lines.append(
                    next_line
                )

            return (

                customer_name,

                ", ".join(
                    address_lines
                )

            )

    return "", ""


# ============================================================
# COMPANY ADDRESS
# ============================================================

def extract_company_address(
    text,
    company
):

    lines = [

        clean_text(line)

        for line
        in text.splitlines()

        if clean_text(line)

    ]

    for index, line in enumerate(
        lines
    ):

        if (

            company
            and
            company.lower()
            in line.lower()

        ):

            address = []

            for next_line in lines[
                index + 1:
                index + 5
            ]:

                lower = next_line.lower()

                if any(
                    word in lower
                    for word in [
                        "invoice",
                        "bill to",
                        "description",
                        "invoice number"
                    ]
                ):

                    break

                address.append(
                    next_line
                )

            return ", ".join(
                address
            )

    return ""


# ============================================================
# ITEMS
# ============================================================

def extract_invoice_items(
    text
):

    lines = [

        clean_text(line)

        for line
        in text.splitlines()

        if clean_text(line)

    ]

    items = []

    started = False

    stop_words = [

        "subtotal",
        "sub total",
        "tax",
        "vat",
        "shipping",
        "discount",
        "total",
        "paid",
        "balance due"

    ]

    for line in lines:

        lower = line.lower()

        if (

            "description" in lower
            and
            (
                "amount" in lower
                or
                "price" in lower
                or
                "qty" in lower
                or
                "quantity" in lower
            )

        ):

            started = True

            continue

        if not started:

            continue

        if any(
            word in lower
            for word in stop_words
        ):

            break

        money_matches = re.findall(

            r"[$S]\s*"
            r"[\d,]+"
            r"(?:\.\d{1,2})?",

            line

        )

        if not money_matches:

            continue

        amount = clean_money(
            money_matches[-1]
        )

        description = line.replace(

            money_matches[-1],

            ""

        )

        description = clean_text(
            description
        )

        description = re.sub(

            r"^\s*\d+\s*[.)_-]?\s*",

            "",

            description

        )

        if description:

            items.append({

                "description":
                    description,

                "amount":
                    amount

            })

    return items


# ============================================================
# COMPLETE INVOICE DATA
# ============================================================

def extract_invoice_data(
    aligned_ocr
):

    company = extract_company(
        aligned_ocr
    )

    company_address = (
        extract_company_address(
            aligned_ocr,
            company
        )
    )

    customer, customer_address = (
        extract_customer(
            aligned_ocr
        )
    )

    invoice_number = (
        extract_invoice_number(
            aligned_ocr
        )
    )

    invoice_date = find_label_value(

        aligned_ocr,

        [
            "Date"
        ]

    )

    due_date = find_label_value(

        aligned_ocr,

        [
            "due_date",
            "due date",
            "Due Date"
        ]

    )

    payment_terms = find_label_value(

        aligned_ocr,

        [
            "Term",
            "Terms",
            "Payment Terms"
        ]

    )

    subtotal = find_money_after_label(

        aligned_ocr,

        [
            "Subtotal",
            "Sub Total"
        ]

    )

    tax = find_money_after_label(

        aligned_ocr,

        [
            "Tax",
            "VAT",
            "Sales Tax"
        ]

    )

    discount = find_money_after_label(

        aligned_ocr,

        [
            "Discount"
        ]

    )

    shipping = find_money_after_label(

        aligned_ocr,

        [
            "Shipping",
            "Handling"
        ]

    )

    grand_total = find_money_after_label(

        aligned_ocr,

        [
            "Grand Total",
            "Total Due",
            "Total"
        ]

    )

    amount_paid = find_money_after_label(

        aligned_ocr,

        [
            "Amount Paid",
            "Paid"
        ]

    )

    balance_due = find_money_after_label(

        aligned_ocr,

        [
            "Balance Due",
            "Amount Due"
        ]

    )

    if "$" in aligned_ocr:

        currency = "$"

    elif "€" in aligned_ocr:

        currency = "€"

    elif "£" in aligned_ocr:

        currency = "£"

    else:

        currency = ""

    items = extract_invoice_items(
        aligned_ocr
    )

    return {

        "company_name":
            company,

        "company_address":
            company_address,

        "customer_name":
            customer,

        "customer_address":
            customer_address,

        "invoice_number":
            invoice_number,

        "invoice_date":
            invoice_date,

        "due_date":
            due_date,

        "payment_terms":
            payment_terms,

        "currency":
            currency,

        "subtotal":
            subtotal,

        "discount":
            discount,

        "tax":
            tax,

        "shipping":
            shipping,

        "grand_total":
            grand_total,

        "amount_paid":
            amount_paid,

        "balance_due":
            balance_due,

        "items":
            items,

        "extracted_date_time":
            datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            )

    }


# ============================================================
# CREATE EXCEL ROW
# ============================================================

def create_excel_row(
    data,
    source_file
):

    item_text = " | ".join(

        [

            (
                f"{item.get('description', '')} "
                f"({item.get('amount', '')})"
            )

            for item in data.get(
                "items",
                []
            )

        ]

    )

    return {

        "Source File":
            source_file,

        "Company Name":
            data.get(
                "company_name",
                ""
            ),

        "Company Address":
            data.get(
                "company_address",
                ""
            ),

        "Customer / Bill To":
            data.get(
                "customer_name",
                ""
            ),

        "Customer Address":
            data.get(
                "customer_address",
                ""
            ),

        "Invoice Number":
            data.get(
                "invoice_number",
                ""
            ),

        "Invoice Date":
            data.get(
                "invoice_date",
                ""
            ),

        "Due Date":
            data.get(
                "due_date",
                ""
            ),

        "Payment Terms":
            data.get(
                "payment_terms",
                ""
            ),

        "Currency":
            data.get(
                "currency",
                ""
            ),

        "Subtotal":
            data.get(
                "subtotal",
                ""
            ),

        "Discount":
            data.get(
                "discount",
                ""
            ),

        "Tax":
            data.get(
                "tax",
                ""
            ),

        "Shipping":
            data.get(
                "shipping",
                ""
            ),

        "Grand Total":
            data.get(
                "grand_total",
                ""
            ),

        "Amount Paid":
            data.get(
                "amount_paid",
                ""
            ),

        "Balance Due":
            data.get(
                "balance_due",
                ""
            ),

        "Invoice Items":
            item_text,

        "Extracted Date Time":
            data.get(
                "extracted_date_time",
                ""
            )

    }


# ============================================================
# ADD ROW
# ============================================================

def add_invoice_row(
    row
):

    invoice_number = clean_text(

        row.get(
            "Invoice Number",
            ""
        )

    )

    source_file = clean_text(

        row.get(
            "Source File",
            ""
        )

    )

    for existing in (
        st.session_state.excel_rows
    ):

        existing_invoice = clean_text(

            existing.get(
                "Invoice Number",
                ""
            )

        )

        existing_source = clean_text(

            existing.get(
                "Source File",
                ""
            )

        )

        # ----------------------------------------------------
        # DUPLICATE INVOICE NUMBER
        # ----------------------------------------------------

        if (

            invoice_number
            and
            existing_invoice
            and
            invoice_number.lower()
            ==
            existing_invoice.lower()

        ):

            return False

        # ----------------------------------------------------
        # DUPLICATE FILE
        # ----------------------------------------------------

        if (

            source_file
            and
            existing_source
            and
            source_file.lower()
            ==
            existing_source.lower()

        ):

            return False

    st.session_state.excel_rows.append(
        row
    )

    return True


# ============================================================
# READ EXISTING EXCEL
# ============================================================

def read_existing_excel(
    uploaded_excel
):

    workbook = load_workbook(
        uploaded_excel,
        data_only=True
    )

    worksheet = workbook.active

    rows = list(
        worksheet.iter_rows(
            values_only=True
        )
    )

    if not rows:

        return []

    headers = [

        str(value)
        if value is not None
        else ""

        for value in rows[0]

    ]

    result = []

    for row in rows[1:]:

        row_dict = {}

        for index, header in enumerate(
            headers
        ):

            if not header:
                continue

            value = ""

            if index < len(row):

                value = row[index]

            if value is None:

                value = ""

            row_dict[header] = value

        if any(
            str(value).strip()
            for value in row_dict.values()
        ):

            result.append(
                row_dict
            )

    return result


# ============================================================
# CREATE EXCEL
# ============================================================

def create_excel_file(
    rows
):

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Invoices"

    if rows:

        headers = list(
            rows[0].keys()
        )

    else:

        headers = [
            "Invoice Number"
        ]

    worksheet.append(
        headers
    )

    for row in rows:

        worksheet.append(

            [

                row.get(
                    header,
                    ""
                )

                for header in headers

            ]

        )

    # --------------------------------------------------------
    # HEADER STYLE
    # --------------------------------------------------------

    for cell in worksheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    worksheet.freeze_panes = "A2"

    # --------------------------------------------------------
    # WIDTH
    # --------------------------------------------------------

    for column in worksheet.columns:

        max_length = 0

        column_letter = (
            get_column_letter(
                column[0].column
            )
        )

        for cell in column:

            try:

                max_length = max(

                    max_length,

                    len(
                        str(
                            cell.value
                        )
                    )

                )

            except:

                pass

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max_length + 2,
            50
        )

    output = io.BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    return output.getvalue()


# ============================================================
# PROCESS ONE INVOICE
# ============================================================

def process_one_invoice(
    uploaded_file
):

    image = None

    try:

        # ----------------------------------------------------
        # READ IMAGE
        # ----------------------------------------------------

        image = Image.open(
            uploaded_file
        ).convert(
            "RGB"
        )

        # ----------------------------------------------------
        # OCR
        # ----------------------------------------------------

        words, boxes = extract_ocr(

            image,

            reader

        )

        if not words:

            return {

                "success": False,

                "error":
                    "No text detected",

                "image": None,

                "ocr": "",

                "data": None

            }

        # ----------------------------------------------------
        # VISUAL LINES
        # ----------------------------------------------------

        lines = create_visual_lines(

            words,

            boxes

        )

        # ----------------------------------------------------
        # 2D OCR
        # ----------------------------------------------------

        aligned_ocr = create_aligned_ocr(
            lines
        )

        if not aligned_ocr.strip():

            return {

                "success": False,

                "error":
                    "2D OCR produced no text",

                "image": None,

                "ocr": "",

                "data": None

            }

        # ----------------------------------------------------
        # EXTRACT DATA
        # ----------------------------------------------------

        data = extract_invoice_data(
            aligned_ocr
        )

        # ----------------------------------------------------
        # SAVE SMALL PREVIEW ONLY
        #
        # Resize image before keeping it in session.
        # This prevents memory explosion with 120 images.
        # ----------------------------------------------------

        preview = image.copy()

        preview.thumbnail(
            (900, 900)
        )

        return {

            "success": True,

            "error": "",

            "image": preview,

            "ocr": aligned_ocr,

            "data": data

        }

    finally:

        # ----------------------------------------------------
        # MEMORY CLEANUP
        # ----------------------------------------------------

        if image is not None:

            image.close()

        gc.collect()

        if torch.cuda.is_available():

            torch.cuda.empty_cache()


# ============================================================
# APPLICATION TITLE
# ============================================================

st.title(
    "📄 Invoice OCR System"
)

st.write(
    "Extract invoice information from one image "
    "or process a complete folder of invoices."
)

st.divider()


# ============================================================
# EXCEL SECTION
# ============================================================

st.subheader(
    "📊 Excel Database"
)

excel_mode = st.radio(

    "Choose Excel mode",

    [
        "🆕 Create New Excel",
        "📂 Continue Existing Excel"
    ],

    horizontal=True

)


# ============================================================
# NEW EXCEL
# ============================================================

if excel_mode == "🆕 Create New Excel":

    if st.button(
        "🆕 Start New Excel",
        use_container_width=True
    ):

        st.session_state.excel_rows = []

        st.session_state.processing_results = []

        st.session_state.failed_files = []

        st.session_state.processed_file_names = set()

        st.session_state.loaded_excel_name = None

        st.session_state.processing_complete = False

        st.success(
            "New Excel database created."
        )


# ============================================================
# EXISTING EXCEL
# ============================================================

else:

    existing_excel = st.file_uploader(

        "Upload your existing Excel file",

        type=["xlsx"],

        key="continue_excel"

    )

    if existing_excel is not None:

        if (

            st.session_state.loaded_excel_name
            != existing_excel.name

        ):

            try:

                existing_rows = (
                    read_existing_excel(
                        existing_excel
                    )
                )

                st.session_state.excel_rows = (
                    existing_rows
                )

                st.session_state.loaded_excel_name = (
                    existing_excel.name
                )

                st.success(

                    f"Loaded "
                    f"{len(existing_rows)} "
                    f"existing invoice records."

                )

            except Exception as e:

                st.error(
                    "Could not read the Excel file."
                )

                st.exception(e)


st.divider()


# ============================================================
# IMAGE UPLOAD
# ============================================================

st.subheader(
    "🖼️ Invoice Upload"
)

upload_mode = st.radio(

    "Choose upload type",

    [
        "🖼️ Single Image",
        "📁 Multiple Images / Folder"
    ],

    horizontal=True

)

uploaded_files = []


# ============================================================
# SINGLE IMAGE
# ============================================================

if upload_mode == "🖼️ Single Image":

    single_file = st.file_uploader(

        "Select one invoice",

        type=[
            "jpg",
            "jpeg",
            "png"
        ],

        accept_multiple_files=False,

        key="single_invoice_upload"

    )

    if single_file is not None:

        uploaded_files = [
            single_file
        ]


# ============================================================
# MULTIPLE / FOLDER
# ============================================================

else:

    multiple_files = st.file_uploader(

        "Select multiple invoices or a folder",

        type=[
            "jpg",
            "jpeg",
            "png"
        ],

        accept_multiple_files="directory",

        key="bulk_invoice_upload"

    )

    if multiple_files:

        uploaded_files = list(
            multiple_files
        )


# ============================================================
# SHOW FILE COUNT
# ============================================================

if uploaded_files:

    st.success(

        f"📁 {len(uploaded_files)} "
        f"invoice image(s) selected."

    )


# ============================================================
# BULK PROCESSING SETTINGS
# ============================================================

if uploaded_files:

    if len(uploaded_files) > 1:

        st.caption(

            f"Processing is done in batches of "
            f"{BATCH_SIZE} invoices to prevent memory "
            f"problems."

        )


# ============================================================
# PROCESS BUTTON
# ============================================================

if uploaded_files:

    process_button = st.button(

        "🔍 Process Invoice(s)",

        type="primary",

        use_container_width=True

    )

    if process_button:

        total_files = len(
            uploaded_files
        )

        successful = 0

        failed = 0

        progress_bar = st.progress(
            0
        )

        status_text = st.empty()

        # ----------------------------------------------------
        # PROCESS IN SMALL BATCHES
        # ----------------------------------------------------

        for batch_start in range(
            0,
            total_files,
            BATCH_SIZE
        ):

            batch_end = min(

                batch_start
                +
                BATCH_SIZE,

                total_files

            )

            batch = uploaded_files[
                batch_start:
                batch_end
            ]

            # ------------------------------------------------
            # PROCESS EACH FILE
            # ------------------------------------------------

            for local_index, uploaded_file in enumerate(
                batch
            ):

                global_index = (
                    batch_start
                    +
                    local_index
                )

                file_name = (
                    uploaded_file.name
                )

                status_text.info(

                    f"Processing "
                    f"{global_index + 1} "
                    f"/ {total_files}: "
                    f"{file_name}"

                )

                # --------------------------------------------
                # DON'T PROCESS SAME FILE TWICE
                # --------------------------------------------

                if file_name in (
                    st.session_state.processed_file_names
                ):

                    progress_bar.progress(

                        int(

                            (
                                global_index + 1
                            )
                            /
                            total_files
                            *
                            100

                        )

                    )

                    continue

                try:

                    result = process_one_invoice(

                        uploaded_file

                    )

                    # ----------------------------------------
                    # SUCCESS
                    # ----------------------------------------

                    if result["success"]:

                        data = result["data"]

                        excel_row = (
                            create_excel_row(

                                data,

                                file_name

                            )
                        )

                        added = add_invoice_row(
                            excel_row
                        )

                        successful += 1

                        st.session_state.processing_results.append({

                            "file_name":
                                file_name,

                            "image":
                                result["image"],

                            "ocr":
                                result["ocr"],

                            "data":
                                data,

                            "added":
                                added

                        })

                    # ----------------------------------------
                    # FAILED
                    # ----------------------------------------

                    else:

                        failed += 1

                        st.session_state.failed_files.append({

                            "file_name":
                                file_name,

                            "error":
                                result["error"]

                        })

                except Exception as e:

                    failed += 1

                    st.session_state.failed_files.append({

                        "file_name":
                            file_name,

                        "error":
                            str(e)

                    })

                finally:

                    st.session_state.processed_file_names.add(
                        file_name
                    )

                    # ----------------------------------------
                    # UPDATE PROGRESS
                    # ----------------------------------------

                    progress_bar.progress(

                        int(

                            (
                                global_index + 1
                            )
                            /
                            total_files
                            *
                            100

                        )

                    )

                    # ----------------------------------------
                    # MEMORY CLEANUP
                    # ----------------------------------------

                    gc.collect()

                    if torch.cuda.is_available():

                        torch.cuda.empty_cache()

            # ------------------------------------------------
            # BATCH CLEANUP
            # ------------------------------------------------

            gc.collect()

            if torch.cuda.is_available():

                torch.cuda.empty_cache()

        progress_bar.progress(
            100
        )

        status_text.success(
            "Finished processing all selected invoices."
        )

        st.session_state.processing_complete = True

        st.success(

            f"Completed: "
            f"{successful} successful, "
            f"{failed} failed."

        )


# ============================================================
# PROCESSING SUMMARY
# ============================================================

if (

    st.session_state.processing_results
    or
    st.session_state.failed_files

):

    st.divider()

    st.subheader(
        "📈 Processing Summary"
    )

    successful_count = len(
        st.session_state.processing_results
    )

    failed_count = len(
        st.session_state.failed_files
    )

    col1, col2, col3 = st.columns(
        3
    )

    with col1:

        st.metric(
            "Successfully Processed",
            successful_count
        )

    with col2:

        st.metric(
            "Failed / No Text",
            failed_count
        )

    with col3:

        st.metric(
            "Excel Records",
            len(
                st.session_state.excel_rows
            )
        )


# ============================================================
# SHOW FAILED FILES
# ============================================================

if st.session_state.failed_files:

    with st.expander(
        "⚠️ Failed Invoices",
        expanded=False
    ):

        failed_df = pd.DataFrame(
            st.session_state.failed_files
        )

        st.dataframe(
            failed_df,
            use_container_width=True
        )

        st.caption(

            "These files did not stop the batch. "
            "The remaining invoices continued processing."

        )


# ============================================================
# SHOW PROCESSED INVOICES
# ============================================================

if st.session_state.processing_results:

    st.divider()

    st.subheader(
        "📄 Extracted Invoices"
    )

    for index, result in enumerate(

        st.session_state.processing_results

    ):

        with st.expander(

            f"{index + 1}. "
            f"{result['file_name']}",

            expanded=(
                index == 0
            )

        ):

            left, right = st.columns(
                2
            )

            # ----------------------------------------------
            # ORIGINAL IMAGE
            # ----------------------------------------------

            with left:

                if result["image"] is not None:

                    st.image(

                        result["image"],

                        use_container_width=True

                    )

            # ----------------------------------------------
            # 2D OCR
            # ----------------------------------------------

            with right:

                st.code(

                    result["ocr"],

                    language=None

                )

            # ----------------------------------------------
            # DATA
            # ----------------------------------------------

            data = result["data"]

            st.subheader(
                "Extracted Data"
            )

            col1, col2 = st.columns(
                2
            )

            with col1:

                st.text_input(

                    "Company Name",

                    value=data.get(
                        "company_name",
                        ""
                    ),

                    key=f"company_{index}"

                )

                st.text_area(

                    "Company Address",

                    value=data.get(
                        "company_address",
                        ""
                    ),

                    key=f"company_address_{index}"

                )

                st.text_input(

                    "Customer / Bill To",

                    value=data.get(
                        "customer_name",
                        ""
                    ),

                    key=f"customer_{index}"

                )

                st.text_area(

                    "Customer Address",

                    value=data.get(
                        "customer_address",
                        ""
                    ),

                    key=f"customer_address_{index}"

                )

                st.text_input(

                    "Invoice Number",

                    value=data.get(
                        "invoice_number",
                        ""
                    ),

                    key=f"invoice_number_{index}"

                )

            with col2:

                st.text_input(

                    "Invoice Date",

                    value=data.get(
                        "invoice_date",
                        ""
                    ),

                    key=f"invoice_date_{index}"

                )

                st.text_input(

                    "Due Date",

                    value=data.get(
                        "due_date",
                        ""
                    ),

                    key=f"due_date_{index}"

                )

                st.text_input(

                    "Payment Terms",

                    value=data.get(
                        "payment_terms",
                        ""
                    ),

                    key=f"payment_terms_{index}"

                )

                st.text_input(

                    "Subtotal",

                    value=data.get(
                        "subtotal",
                        ""
                    ),

                    key=f"subtotal_{index}"

                )

                st.text_input(

                    "Tax",

                    value=data.get(
                        "tax",
                        ""
                    ),

                    key=f"tax_{index}"

                )

                st.text_input(

                    "Grand Total",

                    value=data.get(
                        "grand_total",
                        ""
                    ),

                    key=f"grand_total_{index}"

                )

                st.text_input(

                    "Amount Paid",

                    value=data.get(
                        "amount_paid",
                        ""
                    ),

                    key=f"amount_paid_{index}"

                )

                st.text_input(

                    "Balance Due",

                    value=data.get(
                        "balance_due",
                        ""
                    ),

                    key=f"balance_due_{index}"

                )

            # ----------------------------------------------
            # ITEMS
            # ----------------------------------------------

            if data.get("items"):

                st.subheader(
                    "Invoice Items"
                )

                item_df = pd.DataFrame(
                    data["items"]
                )

                st.dataframe(

                    item_df,

                    use_container_width=True

                )


# ============================================================
# EXCEL DATABASE
# ============================================================

st.divider()

st.subheader(
    "📊 Current Excel Database"
)

st.write(

    f"Total invoice records: "
    f"**{len(st.session_state.excel_rows)}**"

)


# ============================================================
# SHOW EXCEL DATA
# ============================================================

if st.session_state.excel_rows:

    excel_df = pd.DataFrame(
        st.session_state.excel_rows
    )

    st.dataframe(

        excel_df,

        use_container_width=True,

        height=400

    )

    # --------------------------------------------------------
    # CREATE EXCEL
    # --------------------------------------------------------

    excel_bytes = create_excel_file(

        st.session_state.excel_rows

    )

    st.download_button(

        label=(
            "📥 Download Complete Excel"
        ),

        data=excel_bytes,

        file_name=(
            "invoice_database.xlsx"
        ),

        mime=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),

        use_container_width=True

    )


# ============================================================
# CLEAR SESSION
# ============================================================

st.divider()

if st.button(

    "🗑️ Clear Current Session",

    use_container_width=True

):

    st.session_state.excel_rows = []

    st.session_state.processing_results = []

    st.session_state.failed_files = []

    st.session_state.processed_file_names = set()

    st.session_state.loaded_excel_name = None

    st.session_state.processing_complete = False

    gc.collect()

    if torch.cuda.is_available():

        torch.cuda.empty_cache()

    st.rerun()